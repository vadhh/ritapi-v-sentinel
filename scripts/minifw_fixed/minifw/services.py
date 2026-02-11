"""
Service layer untuk operasi MiniFW-AI
"""
import io
import json
import os
import subprocess
from datetime import timedelta
from pathlib import Path
from typing import Dict, List, Optional

from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.utils import timezone


class MiniFWConfig:
    """Handler untuk policy.json configuration"""
    
    POLICY_PATH = "/opt/minifw_ai/config/policy.json"
    
    @classmethod
    def load_policy(cls) -> Dict:
        """Load policy configuration"""
        try:
            with open(cls.POLICY_PATH, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            return {}
        except json.JSONDecodeError:
            return {}
    
    @classmethod
    def save_policy(cls, policy: Dict) -> bool:
        """Save policy configuration"""
        try:
            # Backup existing policy
            if os.path.exists(cls.POLICY_PATH):
                backup_path = f"{cls.POLICY_PATH}.backup"
                with open(cls.POLICY_PATH, 'r') as f:
                    with open(backup_path, 'w') as bf:
                        bf.write(f.read())
            
            # Write new policy
            with open(cls.POLICY_PATH, 'w') as f:
                json.dump(policy, f, indent=2)
            
            return True
        except Exception as e:
            print(f"Error saving policy: {e}")
            return False
    
    @classmethod
    def get_segments(cls) -> Dict:
        """Get segment configuration"""
        policy = cls.load_policy()
        return policy.get('segments', {})
    
    @classmethod
    def get_segment_subnets(cls) -> Dict:
        """Get segment subnet mappings"""
        policy = cls.load_policy()
        return policy.get('segment_subnets', {})
    
    @classmethod
    def update_segments(cls, segments: Dict) -> bool:
        """Update segment thresholds"""
        policy = cls.load_policy()
        policy['segments'] = segments
        return cls.save_policy(policy)
    
    @classmethod
    def update_segment_subnets(cls, subnets: Dict) -> bool:
        """Update segment subnet mappings"""
        policy = cls.load_policy()
        policy['segment_subnets'] = subnets
        return cls.save_policy(policy)
    
    @classmethod
    def get_features(cls) -> Dict:
        """Get feature weights"""
        policy = cls.load_policy()
        return policy.get('features', {})
    
    @classmethod
    def update_features(cls, features: Dict) -> bool:
        """Update feature weights"""
        policy = cls.load_policy()
        policy['features'] = features
        return cls.save_policy(policy)
    
    @classmethod
    def get_enforcement(cls) -> Dict:
        """Get enforcement configuration"""
        policy = cls.load_policy()
        return policy.get('enforcement', {})
    
    @classmethod
    def get_burst(cls) -> Dict:
        """Get burst configuration"""
        policy = cls.load_policy()
        return policy.get('burst', {})


class MiniFWFeeds:
    """Handler untuk feed files (allow/deny lists)"""
    
    FEEDS_DIR = "/opt/minifw_ai/config/feeds"
    
    FEED_FILES = {
        'allow_domains': 'allow_domains.txt',
        'deny_domains': 'deny_domains.txt',
        'deny_ips': 'deny_ips.txt',
        'deny_asn': 'deny_asn.txt',
    }
    
    @classmethod
    def read_feed(cls, feed_name: str) -> List[str]:
        """Read feed file and return list of entries"""
        file_path = os.path.join(cls.FEEDS_DIR, cls.FEED_FILES.get(feed_name, ''))
        
        if not os.path.exists(file_path):
            return []
        
        try:
            with open(file_path, 'r') as f:
                lines = f.readlines()
            
            # Filter out comments and empty lines
            entries = []
            for line in lines:
                line = line.strip()
                if line and not line.startswith('#'):
                    entries.append(line)
            
            return entries
        except Exception as e:
            print(f"Error reading feed {feed_name}: {e}")
            return []
    
    @classmethod
    def write_feed(cls, feed_name: str, entries: List[str]) -> bool:
        """Write entries to feed file"""
        file_path = os.path.join(cls.FEEDS_DIR, cls.FEED_FILES.get(feed_name, ''))
        
        try:
            # Backup existing file
            if os.path.exists(file_path):
                backup_path = f"{file_path}.backup"
                with open(file_path, 'r') as f:
                    with open(backup_path, 'w') as bf:
                        bf.write(f.read())
            
            # Write new entries
            with open(file_path, 'w') as f:
                f.write("# Updated via RITAPI Dashboard\n")
                f.write(f"# Total entries: {len(entries)}\n\n")
                for entry in entries:
                    if entry.strip():
                        f.write(f"{entry.strip()}\n")
            
            return True
        except Exception as e:
            print(f"Error writing feed {feed_name}: {e}")
            return False
    
    @classmethod
    def add_to_feed(cls, feed_name: str, entry: str) -> bool:
        """Add single entry to feed"""
        entries = cls.read_feed(feed_name)
        if entry not in entries:
            entries.append(entry)
            return cls.write_feed(feed_name, entries)
        return True
    
    @classmethod
    def remove_from_feed(cls, feed_name: str, entry: str) -> bool:
        """Remove single entry from feed"""
        entries = cls.read_feed(feed_name)
        if entry in entries:
            entries.remove(entry)
            return cls.write_feed(feed_name, entries)
        return True


class MiniFWService:
    """Handler untuk MiniFW-AI service operations"""
    
    SERVICE_NAME = "minifw-ai"
    
    @classmethod
    def get_status(cls) -> Dict:
        """Get service status"""
        try:
            result = subprocess.run(
                ['systemctl', 'is-active', cls.SERVICE_NAME],
                capture_output=True,
                text=True
            )
            is_active = result.returncode == 0

            result = subprocess.run(
                ['systemctl', 'is-enabled', cls.SERVICE_NAME],
                capture_output=True,
                text=True
            )
            is_enabled = result.returncode == 0

            return {
                'active': is_active,
                'enabled': is_enabled,
                'status': 'running' if is_active else 'stopped'
            }
        except Exception as e:
            return {
                'active': False,
                'enabled': False,
                'status': 'unknown',
                'error': str(e)
            }
    
    @classmethod
    def restart(cls) -> bool:
        """Restart MiniFW-AI service"""
        try:
            subprocess.run(['systemctl', 'restart', cls.SERVICE_NAME], check=True)
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return False

    @classmethod
    def stop(cls) -> bool:
        """Stop MiniFW-AI service"""
        try:
            subprocess.run(['systemctl', 'stop', cls.SERVICE_NAME], check=True)
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return False

    @classmethod
    def start(cls) -> bool:
        """Start MiniFW-AI service"""
        try:
            subprocess.run(['systemctl', 'start', cls.SERVICE_NAME], check=True)
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return False


class MiniFWIPSet:
    """Handler untuk ipset operations"""
    
    IPSET_NAME = "minifw_block_v4"
    
    @classmethod
    def list_blocked_ips(cls) -> List[str]:
        """List all blocked IPs from ipset"""
        try:
            result = subprocess.run(
                ['ipset', 'list', cls.IPSET_NAME],
                capture_output=True,
                text=True,
                check=True
            )

            # Parse ipset output
            ips = []
            in_members = False
            for line in result.stdout.split('\n'):
                if line.startswith('Members:'):
                    in_members = True
                    continue
                if in_members and line.strip():
                    # Format: IP timeout VALUE
                    parts = line.strip().split()
                    if parts:
                        ips.append(parts[0])

            return ips
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return []

    @classmethod
    def add_ip(cls, ip: str, timeout: int = 86400) -> bool:
        """Add IP to block list"""
        try:
            subprocess.run(
                ['ipset', 'add', cls.IPSET_NAME, ip, 'timeout', str(timeout)],
                check=True
            )
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return False

    @classmethod
    def remove_ip(cls, ip: str) -> bool:
        """Remove IP from block list"""
        try:
            subprocess.run(
                ['ipset', 'del', cls.IPSET_NAME, ip],
                check=True
            )
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return False

    @classmethod
    def flush_all(cls) -> bool:
        """Flush all blocked IPs"""
        try:
            subprocess.run(
                ['ipset', 'flush', cls.IPSET_NAME],
                check=True
            )
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, PermissionError, OSError):
            return False


class MiniFWStats:
    """Handler untuk statistics dan monitoring"""
    
    EVENTS_LOG = "/opt/minifw_ai/logs/events.jsonl"
    
    @classmethod
    def get_recent_events(cls, limit: int = 100) -> List[Dict]:
        """Get recent events from JSONL log"""
        if not os.path.exists(cls.EVENTS_LOG):
            return []
        
        events = []
        try:
            with open(cls.EVENTS_LOG, 'r') as f:
                # Read last N lines
                lines = f.readlines()
                for line in lines[-limit:]:
                    try:
                        event = json.loads(line.strip())
                        events.append(event)
                    except json.JSONDecodeError:
                        continue
            
            return events
        except Exception as e:
            print(f"Error reading events: {e}")
            return []
    
    @classmethod
    def get_stats(cls) -> Dict:
        """Get statistics from events"""
        events = cls.get_recent_events(1000)
        
        stats = {
            'total_events': len(events),
            'blocked': 0,
            'monitored': 0,
            'allowed': 0,
            'top_blocked_ips': {},
            'top_blocked_domains': {},
            'by_segment': {}
        }
        
        for event in events:
            action = event.get('action', 'unknown')
            if action == 'block':
                stats['blocked'] += 1
            elif action == 'monitor':
                stats['monitored'] += 1
            elif action == 'allow':
                stats['allowed'] += 1
            
            # Count by IP
            ip = event.get('client_ip', 'unknown')
            if action == 'block':
                stats['top_blocked_ips'][ip] = stats['top_blocked_ips'].get(ip, 0) + 1
            
            # Count by domain
            domain = event.get('domain', 'unknown')
            if action == 'block':
                stats['top_blocked_domains'][domain] = stats['top_blocked_domains'].get(domain, 0) + 1
            
            # Count by segment
            segment = event.get('segment', 'unknown')
            if segment not in stats['by_segment']:
                stats['by_segment'][segment] = {'blocked': 0, 'monitored': 0, 'allowed': 0}
            stats['by_segment'][segment][action] = stats['by_segment'][segment].get(action, 0) + 1
        
        # Sort top IPs and domains
        stats['top_blocked_ips'] = dict(sorted(stats['top_blocked_ips'].items(), key=lambda x: x[1], reverse=True)[:10])
        stats['top_blocked_domains'] = dict(sorted(stats['top_blocked_domains'].items(), key=lambda x: x[1], reverse=True)[:10])
        
        return stats


class AuditService:
    """Service for recording and querying audit logs"""

    @classmethod
    def log_action(cls, request, action, description, severity='info',
                   resource_type=None, resource_id=None,
                   before_value=None, after_value=None):
        """Record a user action in the audit log"""
        try:
            from .models import AuditLog
            user = request.user

            AuditLog.objects.create(
                username=user.username if user.is_authenticated else "anonymous",
                user_role=RBACService.get_user_role(user) if user.is_authenticated else 'anonymous',
                action=action,
                description=description,
                severity=severity,
                resource_type=resource_type,
                resource_id=resource_id,
                ip_address=request.META.get('REMOTE_ADDR'),
                before_value=before_value,
                after_value=after_value,
            )
            return True
        except Exception as e:
            print(f"Error logging audit action: {e}")
            return False

    @classmethod
    def get_logs(cls, limit=50, offset=0, filters=None):
        """Query AuditLog with optional filters, return paginated results."""
        from .models import AuditLog
        qs = AuditLog.objects.all()

        if filters:
            if filters.get('action'):
                qs = qs.filter(action=filters['action'])
            if filters.get('severity'):
                qs = qs.filter(severity=filters['severity'])
            if filters.get('username'):
                qs = qs.filter(username__icontains=filters['username'])
            if filters.get('resource_type'):
                qs = qs.filter(resource_type=filters['resource_type'])
            if filters.get('start_date'):
                qs = qs.filter(timestamp__gte=filters['start_date'])
            if filters.get('end_date'):
                qs = qs.filter(timestamp__lte=filters['end_date'])

        total = qs.count()
        logs = list(
            qs[offset:offset + limit].values(
                'id', 'timestamp', 'username', 'user_role', 'action',
                'severity', 'resource_type', 'resource_id', 'description',
                'ip_address', 'success', 'before_value', 'after_value',
            )
        )
        for log in logs:
            log['timestamp'] = log['timestamp'].isoformat() if log['timestamp'] else None
        return {'total': total, 'logs': logs}

    @classmethod
    def get_statistics(cls, days=7):
        """Aggregate counts by severity for the last N days."""
        from .models import AuditLog
        since = timezone.now() - timedelta(days=days)
        qs = AuditLog.objects.filter(timestamp__gte=since)
        counts = {r['severity']: r['count'] for r in qs.values('severity').annotate(count=Count('id'))}
        return {
            'total': sum(counts.values()),
            'critical': counts.get('critical', 0),
            'warning': counts.get('warning', 0),
            'info': counts.get('info', 0),
        }

    @classmethod
    def export_logs(cls, start_date=None, end_date=None):
        """Serialize audit logs to JSON string for download."""
        from .models import AuditLog
        qs = AuditLog.objects.all()
        if start_date:
            qs = qs.filter(timestamp__gte=start_date)
        if end_date:
            qs = qs.filter(timestamp__lte=end_date)

        logs = list(qs.values(
            'id', 'timestamp', 'username', 'user_role', 'action',
            'severity', 'resource_type', 'resource_id', 'description',
            'ip_address', 'success', 'before_value', 'after_value',
        ))
        for log in logs:
            log['timestamp'] = log['timestamp'].isoformat() if log['timestamp'] else None
        return json.dumps(logs, indent=2, ensure_ascii=False)


class SectorLock:
    """Service for accessing factory-set sector configuration"""
    
    LOCK_FILE = "/opt/minifw_ai/config/sector_lock.json"
    
    @classmethod
    def get_sector(cls) -> str:
        """Get the current locked sector"""
        try:
            if os.path.exists(cls.LOCK_FILE):
                with open(cls.LOCK_FILE, 'r') as f:
                    return json.load(f).get('sector', 'unknown')
        except Exception:
            pass
        
        return os.getenv('MINIFW_SECTOR', 'unknown')
    
    @classmethod
    def get_description(cls) -> str:
        """Get human-readable description of the sector"""
        sector = cls.get_sector()
        descriptions = {
            "hospital": "Healthcare: Prioritizes IoMT device anomaly detection.",
            "school": "Education: Enforces SafeSearch and content filtering.",
            "government": "Government: Strict Geo-IP and long-term audit logging.",
            "finance": "Finance: PCI-DSS compliance and anti-fraud.",
            "legal": "Legal: Data exfiltration detection.",
            "establishment": "Establishment: Balanced protection for SMEs.",
        }
        return descriptions.get(sector, "Standard: Default security policy.")

    @classmethod
    def get_full_config(cls) -> Dict:
        """Return full sector_lock.json content (read-only)."""
        try:
            if os.path.exists(cls.LOCK_FILE):
                with open(cls.LOCK_FILE, 'r') as f:
                    return json.load(f)
        except Exception:
            pass
        return {'sector': os.getenv('MINIFW_SECTOR', 'unknown')}


class DeploymentStateService:
    """Read deployment state from MiniFW-AI's state file for dashboard visibility."""

    STATE_FILE = "/var/log/ritapi/deployment_state.json"

    @classmethod
    def get_state(cls) -> Dict:
        """Read deployment_state.json and return parsed state info."""
        try:
            with open(cls.STATE_FILE, 'r') as f:
                raw = json.load(f)
            protection_state = raw.get('current_protection_state',
                                       raw.get('status', 'BASELINE_PROTECTION'))
            ai_enabled = protection_state == 'AI_ENHANCED_PROTECTION'
            return {
                'protection_state': protection_state,
                'ai_enabled': ai_enabled,
                'last_state_check': raw.get('last_state_check'),
                'service_unavailable': False,
                'unavailable_reason': None,
                'raw': raw,
            }
        except FileNotFoundError:
            return cls._unavailable_state('State file not found')
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            return cls._unavailable_state(f'Invalid state file: {e}')
        except OSError as e:
            return cls._unavailable_state(f'Cannot read state file: {e}')

    @classmethod
    def _unavailable_state(cls, reason: str) -> Dict:
        return {
            'protection_state': 'UNAVAILABLE',
            'ai_enabled': False,
            'last_state_check': None,
            'service_unavailable': True,
            'unavailable_reason': reason,
            'raw': {},
        }

    @classmethod
    def filter_ai_reasons(cls, reasons) -> list:
        """Strip AI-specific reasons (mlp_*, yara_*) from a reasons list."""
        if not isinstance(reasons, list):
            return reasons
        return [r for r in reasons if not r.startswith(('mlp_', 'yara_'))]

    @classmethod
    def filter_event_for_baseline(cls, event: Dict) -> Dict:
        """Remove score and AI reasons from an event dict."""
        filtered = {k: v for k, v in event.items() if k != 'score'}
        if 'reasons' in filtered:
            filtered['reasons'] = cls.filter_ai_reasons(filtered['reasons'])
        return filtered

    @classmethod
    def filter_stats_for_baseline(cls, stats: Dict) -> Dict:
        """Set monitored to None and strip monitored from by_segment."""
        stats = dict(stats)
        stats['monitored'] = None
        if 'by_segment' in stats:
            by_segment = {}
            for seg, counts in stats['by_segment'].items():
                counts = dict(counts)
                counts.pop('monitored', None)
                by_segment[seg] = counts
            stats['by_segment'] = by_segment
        return stats


class MiniFWEventsService:
    """Server-side events processing for DataTables and Excel export."""

    EVENTS_LOG = "/opt/minifw_ai/logs/events.jsonl"

    @classmethod
    def _read_all_events(cls) -> List[Dict]:
        if not os.path.exists(cls.EVENTS_LOG):
            return []
        events = []
        try:
            with open(cls.EVENTS_LOG, 'r') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        try:
                            events.append(json.loads(line))
                        except json.JSONDecodeError:
                            continue
        except Exception:
            pass
        return events

    @classmethod
    def get_events_datatable(cls, draw, start, length, search='', order_col=0, order_dir='desc'):
        """Return DataTables-compatible dict with server-side processing."""
        all_events = cls._read_all_events()
        total = len(all_events)

        # Search filter
        if search:
            search_lower = search.lower()
            all_events = [
                e for e in all_events
                if search_lower in str(e.get('client_ip', '')).lower()
                or search_lower in str(e.get('domain', '')).lower()
                or search_lower in str(e.get('action', '')).lower()
                or search_lower in str(e.get('segment', '')).lower()
            ]
        filtered = len(all_events)

        # Sort
        col_map = {0: 'ts', 1: 'client_ip', 2: 'domain', 3: 'action', 4: 'score', 5: 'segment'}
        sort_key = col_map.get(order_col, 'ts')
        reverse = order_dir == 'desc'
        try:
            all_events.sort(key=lambda e: e.get(sort_key, ''), reverse=reverse)
        except TypeError:
            pass

        # Paginate
        page = all_events[start:start + length]

        return {
            'draw': draw,
            'recordsTotal': total,
            'recordsFiltered': filtered,
            'data': page,
        }

    @classmethod
    def get_event_statistics(cls):
        events = cls._read_all_events()
        stats = {'allowed': 0, 'blocked': 0, 'threats': 0, 'total': len(events)}
        for e in events:
            action = e.get('action', '')
            if action == 'allow':
                stats['allowed'] += 1
            elif action == 'block':
                stats['blocked'] += 1
            elif action in ('deny', 'monitor'):
                stats['threats'] += 1
        return stats

    @classmethod
    def export_events_excel(cls, action_filter=None):
        """Generate openpyxl workbook, return BytesIO."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        events = cls._read_all_events()
        if action_filter and action_filter != 'all':
            events = [e for e in events if e.get('action') == action_filter]

        wb = Workbook()

        # -- Stats sheet --
        ws_stats = wb.active
        ws_stats.title = "Summary"
        header_font = Font(bold=True, size=12)
        ws_stats.append(["MiniFW-AI Security Events Report"])
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats.append([f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_stats.append([f"Total Events: {len(events)}"])
        ws_stats.append([])

        action_counts = {}
        unique_ips = set()
        unique_domains = set()
        for e in events:
            a = e.get('action', 'unknown')
            action_counts[a] = action_counts.get(a, 0) + 1
            unique_ips.add(e.get('client_ip', ''))
            unique_domains.add(e.get('domain', ''))

        ws_stats.append(["Action", "Count"])
        ws_stats['A5'].font = header_font
        ws_stats['B5'].font = header_font
        for action, count in sorted(action_counts.items()):
            ws_stats.append([action, count])
        ws_stats.append([])
        ws_stats.append([f"Unique IPs: {len(unique_ips)}"])
        ws_stats.append([f"Unique Domains: {len(unique_domains)}"])

        # -- Events sheet --
        ws_events = wb.create_sheet("Events")
        columns = ["Timestamp", "Date", "Time", "Client IP", "Domain",
                    "Action", "Score", "Segment", "Reasons"]
        ws_events.append(columns)
        for col_idx in range(1, len(columns) + 1):
            ws_events.cell(row=1, column=col_idx).font = header_font

        action_fills = {
            'allow': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'deny': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'block': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        }

        for e in events:
            ts = e.get('ts', '')
            date_part = ts[:10] if len(ts) >= 10 else ts
            time_part = ts[11:19] if len(ts) >= 19 else ''
            reasons = ', '.join(e.get('reasons', [])) if isinstance(e.get('reasons'), list) else str(e.get('reasons', ''))
            row = [ts, date_part, time_part, e.get('client_ip', ''),
                   e.get('domain', ''), e.get('action', ''),
                   e.get('score', 0), e.get('segment', ''), reasons]
            ws_events.append(row)
            row_num = ws_events.max_row
            fill = action_fills.get(e.get('action'))
            if fill:
                ws_events.cell(row=row_num, column=6).fill = fill

        # Auto-width
        for ws in [ws_stats, ws_events]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or '')))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


class RBACService:
    """Role-based access control helpers."""

    ROLE_HIERARCHY = {
        'SUPER_ADMIN': 5,
        'ADMIN': 4,
        'OPERATOR': 3,
        'AUDITOR': 2,
        'VIEWER': 1,
    }

    @classmethod
    def get_user_role(cls, user):
        if not user or not user.is_authenticated:
            return 'VIEWER'
        profile = getattr(user, 'profile', None)
        if profile is None:
            try:
                from .models import UserProfile
                profile = UserProfile.objects.get(user=user)
            except Exception:
                pass
        if profile:
            return profile.role
        if user.is_superuser:
            return 'SUPER_ADMIN'
        return 'VIEWER'

    @classmethod
    def check_permission(cls, user, required_role):
        user_level = cls.ROLE_HIERARCHY.get(cls.get_user_role(user), 0)
        required_level = cls.ROLE_HIERARCHY.get(required_role, 99)
        return user_level >= required_level

    @classmethod
    def can_modify_policy(cls, user):
        return cls.check_permission(user, 'ADMIN')

    @classmethod
    def can_execute_enforcement(cls, user):
        return cls.check_permission(user, 'OPERATOR')

    @classmethod
    def can_access_audit(cls, user):
        return cls.check_permission(user, 'AUDITOR')

    @classmethod
    def can_export_data(cls, user):
        return cls.check_permission(user, 'AUDITOR')


class UserManagementService:
    """CRUD operations for user accounts + profiles."""

    @classmethod
    def create_user(cls, username, email, password, role, sector,
                    created_by, request=None, **kwargs):
        from .models import UserProfile
        user = User.objects.create_user(username=username, email=email, password=password)
        profile = UserProfile.objects.create(
            user=user,
            role=role,
            sector=sector,
            full_name=kwargs.get('full_name', ''),
            department=kwargs.get('department', ''),
            phone=kwargs.get('phone', ''),
        )
        if request:
            AuditService.log_action(
                request, 'user_created',
                f'Created user {username} with role {role}',
                severity='warning',
                resource_type='user',
                resource_id=str(user.id),
            )
        return user, profile

    @classmethod
    def update_user(cls, user_id, updated_by, request=None, **fields):
        from .models import UserProfile
        user = User.objects.get(id=user_id)
        profile, _ = UserProfile.objects.get_or_create(user=user)

        before = {
            'email': user.email,
            'role': profile.role,
            'sector': profile.sector,
            'full_name': profile.full_name,
            'department': profile.department,
            'phone': profile.phone,
        }

        if 'email' in fields:
            user.email = fields['email']
            user.save(update_fields=['email'])
        profile_fields = ['role', 'sector', 'full_name', 'department', 'phone']
        changed = []
        for f in profile_fields:
            if f in fields:
                setattr(profile, f, fields[f])
                changed.append(f)
        if changed:
            profile.save(update_fields=changed)

        after = {
            'email': user.email,
            'role': profile.role,
            'sector': profile.sector,
            'full_name': profile.full_name,
            'department': profile.department,
            'phone': profile.phone,
        }
        if request:
            AuditService.log_action(
                request, 'user_updated',
                f'Updated user {user.username}: {", ".join(changed or ["email"])}',
                severity='warning',
                resource_type='user',
                resource_id=str(user.id),
                before_value=before,
                after_value=after,
            )
        return user, profile

    @classmethod
    def change_password(cls, user_id, new_password, changed_by, request=None):
        user = User.objects.get(id=user_id)
        user.set_password(new_password)
        user.save(update_fields=['password'])

        from .models import UserProfile
        try:
            profile = user.profile
            profile.last_password_change = timezone.now()
            profile.must_change_password = False
            profile.save(update_fields=['last_password_change', 'must_change_password'])
        except Exception:
            pass

        if request:
            AuditService.log_action(
                request, 'password_changed',
                f'Password changed for user {user.username}',
                severity='warning',
                resource_type='user',
                resource_id=str(user.id),
            )
        return True

    @classmethod
    def delete_user(cls, user_id, deleted_by, request=None):
        user = User.objects.get(id=user_id)
        if user.id == deleted_by:
            raise ValueError("Cannot delete your own account")
        username = user.username
        user.delete()
        if request:
            AuditService.log_action(
                request, 'user_deleted',
                f'Deleted user {username}',
                severity='critical',
                resource_type='user',
                resource_id=str(user_id),
            )
        return True
