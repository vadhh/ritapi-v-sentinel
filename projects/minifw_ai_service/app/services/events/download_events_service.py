import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO


def generate_events_excel_report(action_filter: str = None):
    """
    Generate Excel report from events log with optional action filter
    
    Args:
        action_filter: Filter by action type (allow, deny, block) or None for all
    
    Returns:
        BytesIO object containing the Excel file
    """
    # Read events log
    events = []
    with open('logs/events.jsonl', 'r') as f:
        for line in f:
            try:
                events.append(json.loads(line.strip()))
            except:
                continue
    
    df = pd.DataFrame(events)
    
    if len(df) == 0:
        # Return empty workbook if no events
        wb = Workbook()
        ws = wb.active
        ws.title = "No Events"
        ws['A1'] = "No events found"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # Parse timestamps
    df['ts'] = pd.to_datetime(df['ts'])
    df['date'] = df['ts'].dt.date
    df['time'] = df['ts'].dt.time
    
    # Apply filter if specified
    if action_filter and action_filter.lower() != 'all':
        df = df[df['action'].str.lower() == action_filter.lower()].copy()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cell_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def create_sheet(ws, title, data):
        ws.title = title
        
        # Summary section
        ws['A1'] = 'Security Events Report'
        ws['A1'].font = Font(bold=True, size=14, name='Arial')
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(italic=True, size=10, name='Arial')
        ws.merge_cells('A2:E2')
        
        if action_filter and action_filter.lower() != 'all':
            ws['A3'] = f'Filter: {action_filter.upper()} events only'
            ws['A3'].font = Font(italic=True, size=10, name='Arial', color='FF0000')
            ws.merge_cells('A3:E3')
        
        # Statistics
        row = 4
        ws[f'A{row}'] = 'Total Events:'
        ws[f'B{row}'] = len(data)
        ws[f'A{row}'].font = Font(bold=True, name='Arial')
        
        row += 1
        ws[f'A{row}'] = 'Date Range:'
        ws[f'B{row}'] = f"{data['ts'].min()} to {data['ts'].max()}" if len(data) > 0 else 'N/A'
        ws[f'A{row}'].font = Font(bold=True, name='Arial')
        
        row += 1
        ws[f'A{row}'] = 'Unique Domains:'
        ws[f'B{row}'] = data['domain'].nunique() if len(data) > 0 else 0
        ws[f'A{row}'].font = Font(bold=True, name='Arial')
        
        row += 1
        ws[f'A{row}'] = 'Unique IPs:'
        ws[f'B{row}'] = data['client_ip'].nunique() if len(data) > 0 else 0
        ws[f'A{row}'].font = Font(bold=True, name='Arial')
        
        # Detail section header
        row += 2
        ws[f'A{row}'] = 'Event Details'
        ws[f'A{row}'].font = Font(bold=True, size=12, name='Arial')
        
        # Data headers
        row += 1
        headers = ['Timestamp', 'Date', 'Time', 'Client IP', 'Domain', 'Action', 'Score', 'Segment', 'Reasons']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for idx, event in data.iterrows():
            row += 1
            ws.cell(row=row, column=1, value=str(event['ts']))
            ws.cell(row=row, column=2, value=str(event['date']))
            ws.cell(row=row, column=3, value=str(event['time']))
            ws.cell(row=row, column=4, value=event['client_ip'])
            ws.cell(row=row, column=5, value=event['domain'])
            ws.cell(row=row, column=6, value=event['action'])
            ws.cell(row=row, column=7, value=event['score'])
            ws.cell(row=row, column=8, value=event['segment'])
            ws.cell(row=row, column=9, value=', '.join(event['reasons']) if event['reasons'] else '')
            
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                if col == 6:  # Action column
                    if event['action'] == 'allow':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif event['action'] == 'deny':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    elif event['action'] == 'block':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 50
    
    # Create main data sheet
    create_sheet(wb.create_sheet('Events'), 'Events', df)
    
    # Statistics sheet
    all_df = pd.DataFrame(events)
    if len(all_df) > 0:
        all_df['ts'] = pd.to_datetime(all_df['ts'])
        
        stats_ws = wb.create_sheet('Statistics', 0)
        stats_ws['A1'] = 'Event Statistics'
        stats_ws['A1'].font = Font(bold=True, size=14, name='Arial')
        stats_ws.merge_cells('A1:D1')
        
        row = 3
        stats_ws[f'A{row}'] = 'Action Type'
        stats_ws[f'B{row}'] = 'Count'
        stats_ws[f'C{row}'] = 'Percentage'
        stats_ws[f'D{row}'] = 'Unique Domains'
        for col in range(1, 5):
            cell = stats_ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        actions = all_df['action'].unique()
        total_events = len(all_df)
        for action in sorted(actions):
            row += 1
            count = len(all_df[all_df['action'] == action])
            percentage = (count / total_events * 100) if total_events > 0 else 0
            unique_domains = all_df[all_df['action'] == action]['domain'].nunique()
            
            stats_ws.cell(row=row, column=1, value=action.capitalize())
            stats_ws.cell(row=row, column=2, value=count)
            stats_ws.cell(row=row, column=3, value=f'{percentage:.2f}%')
            stats_ws.cell(row=row, column=4, value=unique_domains)
            
            for col in range(1, 5):
                cell = stats_ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                if col == 2:
                    cell.alignment = Alignment(horizontal='right')
        
        # Total row
        row += 1
        stats_ws.cell(row=row, column=1, value='TOTAL')
        stats_ws.cell(row=row, column=2, value=total_events)
        stats_ws.cell(row=row, column=3, value='100.00%')
        stats_ws.cell(row=row, column=4, value=all_df['domain'].nunique())
        for col in range(1, 5):
            cell = stats_ws.cell(row=row, column=col)
            cell.font = Font(bold=True, name='Arial')
            cell.border = border
            if col == 2:
                cell.alignment = Alignment(horizontal='right')
        
        # Top domains
        row += 3
        stats_ws[f'A{row}'] = 'Top 20 Domains'
        stats_ws[f'A{row}'].font = Font(bold=True, size=12, name='Arial')
        
        row += 1
        stats_ws[f'A{row}'] = 'Domain'
        stats_ws[f'B{row}'] = 'Count'
        stats_ws[f'C{row}'] = 'Allow'
        stats_ws[f'D{row}'] = 'Deny/Block'
        for col in range(1, 5):
            cell = stats_ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        top_domains = all_df['domain'].value_counts().head(20)
        for domain, count in top_domains.items():
            row += 1
            allow_count = len(all_df[(all_df['domain'] == domain) & (all_df['action'] == 'allow')])
            deny_count = len(all_df[(all_df['domain'] == domain) & (all_df['action'] != 'allow')])
            
            stats_ws.cell(row=row, column=1, value=domain)
            stats_ws.cell(row=row, column=2, value=count)
            stats_ws.cell(row=row, column=3, value=allow_count)
            stats_ws.cell(row=row, column=4, value=deny_count)
            
            for col in range(1, 5):
                cell = stats_ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                if col in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='right')
        
        stats_ws.column_dimensions['A'].width = 50
        stats_ws.column_dimensions['B'].width = 12
        stats_ws.column_dimensions['C'].width = 12
        stats_ws.column_dimensions['D'].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output