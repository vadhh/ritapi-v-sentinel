import io
import json
import os
import tempfile
from unittest.mock import patch, MagicMock

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase, RequestFactory, Client

from .services import AuditService, DeploymentStateService, MiniFWEventsService


class DeploymentStateServiceTest(TestCase):
    """Unit tests for DeploymentStateService."""

    def _write_state_file(self, tmpdir, data):
        path = os.path.join(tmpdir, "deployment_state.json")
        with open(path, "w") as f:
            json.dump(data, f)
        return path

    def test_get_state_baseline(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write_state_file(
                tmpdir,
                {
                    "current_protection_state": "BASELINE_PROTECTION",
                },
            )
            with patch.object(DeploymentStateService, "STATE_FILE", path):
                state = DeploymentStateService.get_state()
        self.assertEqual(state["protection_state"], "BASELINE_PROTECTION")
        self.assertFalse(state["ai_enabled"])
        self.assertFalse(state["service_unavailable"])

    def test_get_state_enhanced(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write_state_file(
                tmpdir,
                {
                    "current_protection_state": "AI_ENHANCED_PROTECTION",
                },
            )
            with patch.object(DeploymentStateService, "STATE_FILE", path):
                state = DeploymentStateService.get_state()
        self.assertEqual(state["protection_state"], "AI_ENHANCED_PROTECTION")
        self.assertTrue(state["ai_enabled"])
        self.assertFalse(state["service_unavailable"])

    def test_get_state_missing_file(self):
        with patch.object(
            DeploymentStateService, "STATE_FILE", "/nonexistent/path.json"
        ):
            state = DeploymentStateService.get_state()
        self.assertEqual(state["protection_state"], "UNAVAILABLE")
        self.assertFalse(state["ai_enabled"])
        self.assertTrue(state["service_unavailable"])
        self.assertIn("not found", state["unavailable_reason"])

    def test_get_state_corrupt_json(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "deployment_state.json")
            with open(path, "w") as f:
                f.write("{invalid json!!!")
            with patch.object(DeploymentStateService, "STATE_FILE", path):
                state = DeploymentStateService.get_state()
        self.assertEqual(state["protection_state"], "UNAVAILABLE")
        self.assertTrue(state["service_unavailable"])

    def test_get_state_fallback_status_field(self):
        """Falls back to 'status' key when 'current_protection_state' missing."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = self._write_state_file(
                tmpdir,
                {
                    "status": "AI_ENHANCED_PROTECTION",
                },
            )
            with patch.object(DeploymentStateService, "STATE_FILE", path):
                state = DeploymentStateService.get_state()
        self.assertTrue(state["ai_enabled"])

    def test_filter_ai_reasons(self):
        reasons = ["feed_deny", "mlp_anomaly", "burst_flood", "yara_match", "mlp_high"]
        filtered = DeploymentStateService.filter_ai_reasons(reasons)
        self.assertEqual(filtered, ["feed_deny", "burst_flood"])

    def test_filter_ai_reasons_non_list(self):
        self.assertEqual(DeploymentStateService.filter_ai_reasons("string"), "string")

    def test_filter_event_baseline_removes_score(self):
        event = {
            "ts": "2026-02-11T10:00:00",
            "client_ip": "10.0.0.1",
            "domain": "example.com",
            "action": "block",
            "score": 85,
            "segment": "office",
            "reasons": ["feed_deny", "mlp_anomaly"],
        }
        filtered = DeploymentStateService.filter_event_for_baseline(event)
        self.assertNotIn("score", filtered)
        self.assertEqual(filtered["reasons"], ["feed_deny"])
        self.assertEqual(filtered["client_ip"], "10.0.0.1")

    def test_filter_stats_baseline_hides_monitored(self):
        stats = {
            "total_events": 100,
            "blocked": 40,
            "monitored": 20,
            "allowed": 40,
            "top_blocked_ips": {"10.0.0.1": 5},
            "top_blocked_domains": {"bad.com": 3},
            "by_segment": {
                "office": {"blocked": 10, "monitored": 5, "allowed": 15},
                "lab": {"blocked": 2, "monitored": 1, "allowed": 8},
            },
        }
        filtered = DeploymentStateService.filter_stats_for_baseline(stats)
        self.assertIsNone(filtered["monitored"])
        self.assertNotIn("monitored", filtered["by_segment"]["office"])
        self.assertNotIn("monitored", filtered["by_segment"]["lab"])
        # Other stats preserved
        self.assertEqual(filtered["blocked"], 40)
        self.assertEqual(filtered["top_blocked_ips"], {"10.0.0.1": 5})


def _mock_deployment_state(ai_enabled):
    """Helper to create a mock deployment state."""
    if ai_enabled:
        return {
            "protection_state": "AI_ENHANCED_PROTECTION",
            "ai_enabled": True,
            "last_state_check": None,
            "service_unavailable": False,
            "unavailable_reason": None,
            "raw": {},
        }
    return {
        "protection_state": "BASELINE_PROTECTION",
        "ai_enabled": False,
        "last_state_check": None,
        "service_unavailable": False,
        "unavailable_reason": None,
        "raw": {},
    }


MOCK_STATS = {
    "total_events": 10,
    "blocked": 4,
    "monitored": 3,
    "allowed": 3,
    "top_blocked_ips": {"10.0.0.1": 2},
    "top_blocked_domains": {"bad.com": 1},
    "by_segment": {
        "office": {"blocked": 2, "monitored": 1, "allowed": 2},
    },
}

MOCK_EVENTS = [
    {
        "ts": "2026-02-11T10:00:00",
        "client_ip": "10.0.0.1",
        "domain": "bad.com",
        "action": "block",
        "score": 85,
        "segment": "office",
        "reasons": ["feed_deny", "mlp_anomaly"],
    },
]


class DashboardViewBaselineTest(TestCase):
    """Test dashboard view hides AI metrics in BASELINE mode."""

    def setUp(self):
        from .models import UserProfile

        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )
        UserProfile.objects.create(user=self.user, role="ADMIN", sector="ESTABLISHMENT")
        self.client = Client()
        self.client.login(username="testuser", password="testpass123")

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    @patch("minifw.views.MiniFWStats.get_stats", return_value=MOCK_STATS.copy())
    @patch(
        "minifw.views.MiniFWStats.get_recent_events", return_value=MOCK_EVENTS.copy()
    )
    @patch("minifw.views.MiniFWIPSet.list_blocked_ips", return_value=[])
    @patch(
        "minifw.views.MiniFWService.get_status",
        return_value={"active": True, "enabled": True, "status": "running"},
    )
    @patch("minifw.views.SectorLock.get_sector", return_value="establishment")
    @patch("minifw.views.SectorLock.get_description", return_value="Standard")
    def test_dashboard_baseline_hides_monitored(self, *mocks):
        response = self.client.get("/ops/minifw/dashboard/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertNotIn("Monitored", content)
        self.assertNotIn("Score", content)
        self.assertIn("Baseline Protection", content)

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    @patch("minifw.views.MiniFWStats.get_stats", return_value=MOCK_STATS.copy())
    @patch(
        "minifw.views.MiniFWStats.get_recent_events", return_value=MOCK_EVENTS.copy()
    )
    @patch("minifw.views.MiniFWIPSet.list_blocked_ips", return_value=[])
    @patch(
        "minifw.views.MiniFWService.get_status",
        return_value={"active": True, "enabled": True, "status": "running"},
    )
    @patch("minifw.views.SectorLock.get_sector", return_value="establishment")
    @patch("minifw.views.SectorLock.get_description", return_value="Standard")
    def test_golden_rule_no_ai_terms_in_baseline_html(self, *mocks):
        """Golden Rule: baseline HTML must not contain AI-specific terms or scores."""
        response = self.client.get("/ops/minifw/dashboard/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertNotIn("mlp_anomaly", content)
        self.assertNotIn("yara_match", content)
        # The mock score value 85 should not appear in baseline output
        self.assertNotIn(">85<", content)

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(True),
    )
    @patch("minifw.views.MiniFWStats.get_stats", return_value=MOCK_STATS.copy())
    @patch(
        "minifw.views.MiniFWStats.get_recent_events", return_value=MOCK_EVENTS.copy()
    )
    @patch("minifw.views.MiniFWIPSet.list_blocked_ips", return_value=[])
    @patch(
        "minifw.views.MiniFWService.get_status",
        return_value={"active": True, "enabled": True, "status": "running"},
    )
    @patch("minifw.views.SectorLock.get_sector", return_value="establishment")
    @patch("minifw.views.SectorLock.get_description", return_value="Standard")
    def test_dashboard_enhanced_shows_all(self, *mocks):
        response = self.client.get("/ops/minifw/dashboard/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("Monitored", content)
        self.assertIn("Score", content)
        self.assertIn("AI-Enhanced Protection", content)


class ApiStatsBaselineTest(TestCase):
    """Test API endpoints filter for baseline mode."""

    def setUp(self):
        from .models import UserProfile

        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )
        UserProfile.objects.create(user=self.user, role="ADMIN", sector="ESTABLISHMENT")
        self.client = Client()
        self.client.login(username="testuser", password="testpass123")

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    @patch("minifw.views.MiniFWStats.get_stats", return_value=MOCK_STATS.copy())
    def test_api_stats_baseline_filtered(self, *mocks):
        response = self.client.get("/ops/minifw/api/stats/")
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertIsNone(data["monitored"])

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    @patch(
        "minifw.views.MiniFWStats.get_recent_events", return_value=MOCK_EVENTS.copy()
    )
    def test_api_events_baseline_no_score(self, *mocks):
        response = self.client.get("/ops/minifw/api/events/")
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        for event in data["events"]:
            self.assertNotIn("score", event)
            for reason in event.get("reasons", []):
                self.assertFalse(reason.startswith("mlp_"))
                self.assertFalse(reason.startswith("yara_"))

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    @patch(
        "minifw.views.MiniFWStats.get_recent_events", return_value=MOCK_EVENTS.copy()
    )
    def test_golden_rule_api_events_no_ai_reasons(self, *mocks):
        """Golden Rule: API events in baseline must have no mlp_*/yara_* reasons."""
        response = self.client.get("/ops/minifw/api/events/")
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        for event in data["events"]:
            reasons = event.get("reasons", [])
            ai_reasons = [
                r for r in reasons if r.startswith("mlp_") or r.startswith("yara_")
            ]
            self.assertEqual(
                ai_reasons,
                [],
                f"Baseline event should have no AI reasons, found: {ai_reasons}",
            )


class AuditSanitizeValueTest(TestCase):
    """Tests for AuditService._sanitize_value()."""

    def test_sanitize_value_redacts_secrets(self):
        data = {
            "username": "admin",
            "password": "hunter2",
            "api_key": "sk-12345",
            "auth_header": "Bearer xxx",
        }
        result = AuditService._sanitize_value(data)
        self.assertEqual(result["username"], "admin")
        self.assertEqual(result["password"], "***REDACTED***")
        self.assertEqual(result["api_key"], "***REDACTED***")
        self.assertEqual(result["auth_header"], "***REDACTED***")

    def test_sanitize_value_preserves_safe_keys(self):
        data = {
            "ip_address": "10.0.0.1",
            "action": "block",
            "role": "ADMIN",
            "email": "user@example.com",
        }
        result = AuditService._sanitize_value(data)
        self.assertEqual(result, data)

    def test_sanitize_value_nested(self):
        data = {
            "user": {
                "name": "admin",
                "auth_config": {
                    "password": "secret123",
                    "token": "jwt-xxx",
                    "host": "localhost",
                },
            },
            "items": [
                {"secret_key": "abc", "value": 42},
                "plain_string",
            ],
        }
        result = AuditService._sanitize_value(data)
        self.assertEqual(result["user"]["name"], "admin")
        # auth_config doesn't match SENSITIVE_KEYS, so recurse into it
        self.assertEqual(result["user"]["auth_config"]["password"], "***REDACTED***")
        self.assertEqual(result["user"]["auth_config"]["token"], "***REDACTED***")
        self.assertEqual(result["user"]["auth_config"]["host"], "localhost")
        # 'credentials' key itself matches, so entire value is redacted
        cred_data = {"credentials": {"password": "x"}}
        self.assertEqual(
            AuditService._sanitize_value(cred_data)["credentials"], "***REDACTED***"
        )
        # list items
        self.assertEqual(result["items"][0]["secret_key"], "***REDACTED***")
        self.assertEqual(result["items"][0]["value"], 42)
        self.assertEqual(result["items"][1], "plain_string")

    def test_sanitize_value_none_and_primitives(self):
        self.assertIsNone(AuditService._sanitize_value(None))
        self.assertEqual(AuditService._sanitize_value(42), 42)
        self.assertEqual(AuditService._sanitize_value("hello"), "hello")


class ExportEventsBaselineTest(TestCase):
    """Test Excel export omits Score in baseline mode."""

    @patch.object(
        MiniFWEventsService,
        "_read_all_events",
        return_value=[
            {
                "ts": "2026-02-11T10:00:00",
                "client_ip": "10.0.0.1",
                "domain": "bad.com",
                "action": "block",
                "score": 85,
                "segment": "office",
                "reasons": ["feed_deny", "mlp_anomaly"],
            },
        ],
    )
    @patch.object(
        DeploymentStateService, "get_state", return_value=_mock_deployment_state(False)
    )
    def test_export_events_baseline_omits_score(self, *mocks):
        buf = MiniFWEventsService.export_events_excel(ai_enabled=False)
        self.assertIsInstance(buf, io.BytesIO)
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        ws = wb["Events"]
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("Score", headers)
        self.assertIn("Action", headers)
        self.assertIn("Segment", headers)
        # Check that AI reasons are filtered from data
        reasons_col = headers.index("Reasons") + 1
        reasons_val = ws.cell(row=2, column=reasons_col).value
        self.assertNotIn("mlp_", reasons_val)

    @patch.object(
        MiniFWEventsService,
        "_read_all_events",
        return_value=[
            {
                "ts": "2026-02-11T10:00:00",
                "client_ip": "10.0.0.1",
                "domain": "bad.com",
                "action": "block",
                "score": 85,
                "segment": "office",
                "reasons": ["feed_deny", "mlp_anomaly"],
            },
        ],
    )
    @patch.object(
        DeploymentStateService, "get_state", return_value=_mock_deployment_state(True)
    )
    def test_export_events_enhanced_includes_score(self, *mocks):
        buf = MiniFWEventsService.export_events_excel(ai_enabled=True)
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        ws = wb["Events"]
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Score", headers)


class DeploymentStateAPITest(TestCase):
    """Test deployment state API endpoint."""

    def setUp(self):
        from .models import UserProfile

        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )
        UserProfile.objects.create(
            user=self.user, role="VIEWER", sector="ESTABLISHMENT"
        )
        self.client = Client()
        self.client.login(username="testuser", password="testpass123")

    @patch(
        "minifw.views.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    def test_api_deployment_state_endpoint(self, *mocks):
        response = self.client.get("/ops/minifw/api/deployment-state/")
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertIn("protection_state", data)
        self.assertIn("ai_enabled", data)
        self.assertIn("service_unavailable", data)
        self.assertNotIn("raw", data)

    def test_api_deployment_state_requires_login(self):
        self.client.logout()
        response = self.client.get("/ops/minifw/api/deployment-state/")
        self.assertEqual(response.status_code, 302)


class ManagementCommandTest(TestCase):
    """Test minifw_status management command."""

    @patch(
        "minifw.services.DeploymentStateService.get_state",
        return_value=_mock_deployment_state(False),
    )
    @patch(
        "minifw.services.MiniFWService.get_status",
        return_value={"active": False, "enabled": False, "status": "stopped"},
    )
    @patch("minifw.services.SectorLock.get_sector", return_value="establishment")
    @patch("minifw.services.SectorLock.get_description", return_value="Standard")
    def test_management_command_runs(self, *mocks):
        out = io.StringIO()
        call_command("minifw_status", stdout=out)
        output = out.getvalue()
        self.assertIn("BASELINE_PROTECTION", output)
        self.assertIn("Inactive", output)
